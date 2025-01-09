# export const generateFinalDipData = (
#   differenceList: number[],
#   division: number
# ): number[] => {
#   let finalDipData: number[] = [0];
#   for (let i = 0; i < differenceList.length; i++) {
#     let lastVal = finalDipData.length
#       ? finalDipData[finalDipData.length - 1]
#       : 0;
#     for (let j = 0; j < differenceList[i]; j++) {
#       finalDipData.push(
#         lastVal + Math.round((division / differenceList[i]) * (j + 1))
#       );
#     }
#   }
#   return finalDipData;
# };

# dip_raw=input("Enter Dip Values seperated by comma: ")
# dip_values = dip_raw.split(",")
# dip_values = [int(item) for item in dip_values]
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

zero_datum = 7800

dip_values = [38,75,150,304,455,682,910,1138,1154,1388,1692,1920,2148,2376,2604,2832,3060,3288,3515,3667,3819,4047,4275,4502,4731,4959,5187,5414,5567,5718,5871,6022,6175,6327,6479,6631,6782,6859,6935,7011,7086,7162,7238,7313,7390,7465,7510,7586,7663,7738,7815,7891,7966,8041,8071,8102,8131,8161,8176,8192,8207]
divisions = [5,5,10,10,15,15,15,1,20,15,15,15,15,15,15,15,15,15,10,10,15,15,15,15,15,15,15,10,10,10,10,10,10,10,10,10,5,5,5,5,5,5,5,5,5,3,5,5,5,5,5,5,5,2,2,2,2,1,1,1,1]
divisions = [item*1000 for item in divisions]
dip_values.insert(0,0)
diff_values = []
for i in range(len(dip_values)-1):
    diff_values.append(dip_values[i+1]-dip_values[i])
final_dips = [zero_datum]
for i in range(len(diff_values)):
    last_val = final_dips[-1] if len(final_dips)>0 else 0
    for j in range(diff_values[i]):
        final_dips.append(last_val+round((divisions[i]/diff_values[i])*(j+1)))

ws['A1'].value = 'Dip in mm'
ws['B1'].value = 'Quantity in Liter'

for k in range(len(final_dips)):
    ws[f'A{k+2}'].value = k
    ws[f'B{k+2}'].value = final_dips[k]





wb.save('st38.xlsx')

