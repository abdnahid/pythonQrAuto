from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from openpyxl import Workbook

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("http://localhost:3000/test")
all_nothi = driver.find_elements(By.CLASS_NAME,"nothi_subject")


wb = Workbook()
ws = wb.active
ws['A1']='নথি বিষয়ঃ'
ws['B1']='নথি নম্বরঃ'
current_row = 2
for elem in all_nothi:
    split_elem = elem.text.split('\n')
    sub = split_elem[1].strip('নথি বিষয়ঃ')
    nothi_num = split_elem[2].strip('নথি নম্বরঃ')
    ws.cell(row=current_row,column=1,value=sub)
    ws.cell(row=current_row,column=2,value=nothi_num)
    current_row+=1


wb.save('report.xlsx')
